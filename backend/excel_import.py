"""
Excel Import Router
Handles employee imports from uploaded Excel files and workspace workbooks.
"""

import logging
import os
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

import models
from auth import get_current_user
from database import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/excel", tags=["excel"])


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    normalized = re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()
    aliases = {
        "file code": "file_code",
        "employee code": "file_code",
        "staff no": "file_code",
        "staff number": "file_code",
        "employee id": "file_code",
        "employeeid": "file_code",
        "full name": "full_name",
        "employee name": "full_name",
        "staff name": "full_name",
        "name": "full_name",
        "project": "project",
        "department": "project",
        "status": "status",
        "staff status": "status",
        "employment status": "status",
        "position": "position",
        "designation": "position",
        "job title": "position",
        "role": "position",
        "contact": "contact_number",
        "contact number": "contact_number",
        "contact no": "contact_number",
        "contact no ": "contact_number",
        "phone": "contact_number",
        "telephone": "contact_number",
        "mobile": "contact_number",
        "location": "location",
        "site": "location",
        "office": "location",
        "locker": "locker",
        "date of appointment": "date_of_appointment",
        "appointment date": "date_of_appointment",
        "contract start": "contract_start",
        "contract end": "contract_end",
        "contract review date": "contract_review_date",
        "review date": "contract_review_date",
        "probation end": "probation_end",
        "notice period": "notice_period",
        "employment type": "employment_type",
        "contract type": "employment_type",
        "photo": "photo_url",
        "photo url": "photo_url",
        "profile image": "photo_url",
        "profile photo": "photo_url",
        "avatar": "photo_url",
    }

    if normalized in aliases:
        return aliases[normalized]

    if normalized.startswith("application") or "application" in normalized:
        if "resume" in normalized or "cv" in normalized:
            return "missing_app_resume"
    if "appointment letter" in normalized or "job description" in normalized or "appointment" in normalized:
        return "missing_appointment_letter"
    if "academic" in normalized or "academic docs" in normalized:
        return "missing_academic_docs"
    if "recruitment" in normalized and ("note" in normalized or "report" in normalized or "result" in normalized):
        return "missing_recruitment_notes"
    if "staff id" in normalized or "employee data form" in normalized:
        return "missing_staff_id_form"
    if "performance" in normalized and "appraisal" in normalized:
        return "missing_performance_appraisals"
    if "national id" in normalized or "driving permit" in normalized:
        return "missing_national_id"
    if "policy" in normalized and "declaration" in normalized:
        return "missing_policy_declaration"
    if "end of contract" in normalized and "notice" in normalized:
        return "missing_end_of_contract_notice"

    return normalized


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    if text in {"yes", "true", "1", "x", "y", "complete", "done", "filled", "available"}:
        return True
    if text in {"no", "false", "0", "absent", "pending", "n/a", "na", "none"}:
        return False
    return False


def _parse_missing_flag(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    if text in {"missing", "not present", "absent", "pending", "required", "no", "false", "0", "n/a", "na", "none"}:
        return True
    return False


def _map_row_data_to_employee_fields(row_data: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for raw_key, value in row_data.items():
        if raw_key is None:
            continue
        field_name = _normalize_header(raw_key)
        if not field_name:
            continue
        if field_name in {
            "missing_app_resume",
            "missing_appointment_letter",
            "missing_academic_docs",
            "missing_recruitment_notes",
            "missing_staff_id_form",
            "missing_performance_appraisals",
            "missing_national_id",
            "missing_policy_declaration",
            "missing_end_of_contract_notice",
        }:
            mapped[field_name] = _parse_missing_flag(value)
            continue
        if field_name in {"date_of_appointment", "contract_start", "contract_end", "contract_review_date", "probation_end"}:
            parsed_date = _parse_date(value)
            if parsed_date is not None:
                mapped[field_name] = parsed_date
            continue
        if field_name in {"file_code", "full_name", "project", "status", "position", "contact_number", "location", "locker", "employment_type", "notice_period", "photo_url"}:
            if value is None:
                continue
            text_value = str(value).strip()
            if text_value:
                mapped[field_name] = text_value

    if not mapped.get("status") and mapped.get("project"):
        project_text = str(mapped.get("project", "")).upper()
        if "EXITED" in project_text or "EXIT" in project_text:
            mapped["status"] = "Exited"
        elif "RECESS" in project_text or "ON RECESS" in project_text:
            mapped["status"] = "On Recess"
        else:
            mapped["status"] = "Active"
    elif not mapped.get("status"):
        mapped["status"] = "Active"

    return mapped


def _looks_like_header_row(row: list[Any]) -> bool:
    normalized_values = [_normalize_header(cell) for cell in row]
    if not any(normalized_values):
        return False
    known_tokens = {
        "file_code",
        "full_name",
        "project",
        "status",
        "position",
        "contact_number",
        "location",
        "locker",
        "date_of_appointment",
        "contract_start",
        "contract_end",
        "contract_review_date",
        "probation_end",
        "notice_period",
        "employment_type",
        "photo_url",
        "missing_app_resume",
        "missing_appointment_letter",
        "missing_academic_docs",
        "missing_recruitment_notes",
        "missing_staff_id_form",
        "missing_performance_appraisals",
        "missing_national_id",
        "missing_policy_declaration",
        "missing_end_of_contract_notice",
        "file",
        "contact",
        "name",
        "employee",
        "appointment",
        "recruitment",
        "performance",
        "national",
    }
    return any(token in known_tokens or "application" in token or "resume" in token for token in normalized_values)


def _read_workbook_rows(workbook: openpyxl.Workbook) -> list[tuple[str, list[dict[str, Any]]]]:
    workbook_rows: list[tuple[str, list[dict[str, Any]]]] = []
    for sheet in workbook.worksheets:
        header_row_index = None
        headers: list[str] = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            if _looks_like_header_row(list(row)):
                header_row_index = row_idx
                headers = [_normalize_header(cell) for cell in row]
                break
        if header_row_index is None or not headers:
            continue
        sheet_rows = []
        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    row_data[header] = row[idx]
            if row_data:
                sheet_rows.append(_map_row_data_to_employee_fields(row_data))
        if sheet_rows:
            workbook_rows.append((sheet.title, sheet_rows))
    return workbook_rows


def _upsert_employee_from_mapping(db: Session, employee_data: dict[str, Any]) -> tuple[str, str]:
    file_code = employee_data.get("file_code")
    if not file_code or not str(file_code).strip():
        raise ValueError("Missing employee file code")

    existing = db.query(models.Employee).filter(models.Employee.file_code == str(file_code)).first()
    if existing:
        for field_name, value in employee_data.items():
            if field_name in {"file_code"}:
                continue
            if hasattr(existing, field_name) and value is not None:
                setattr(existing, field_name, value)
        existing.status = existing.status or "Active"
        db.add(existing)
        return str(file_code), "updated"

    employee = models.Employee(**employee_data)
    db.add(employee)
    return str(file_code), "created"


@router.post("/import-employees")
async def import_employees_from_excel(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Bulk import employees from an uploaded Excel workbook."""
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can import employees")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        workbook_rows = _read_workbook_rows(workbook)
        imported_employees = []
        errors = []

        for sheet_name, rows in workbook_rows:
            for row_data in rows:
                try:
                    if not row_data.get("file_code"):
                        continue
                    file_code, status = _upsert_employee_from_mapping(db, row_data)
                    imported_employees.append({"file_code": file_code, "status": status, "sheet": sheet_name})
                except Exception as exc:  # pragma: no cover - defensive logging
                    errors.append(f"{sheet_name}: {exc}")
                    logger.error("Error importing row from sheet %s: %s", sheet_name, exc)

        db.commit()

        return {
            "success": True,
            "message": f"Imported {len(imported_employees)} employees from {len(workbook_rows)} sheets",
            "imported": imported_employees,
            "errors": errors,
            "sheets_processed": [name for name, _ in workbook_rows],
        }
    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(status_code=400, detail="Invalid Excel file format")
    except Exception as exc:
        logger.error("Error importing Excel file: %s", exc)
        raise HTTPException(status_code=400, detail=f"Error importing file: {exc}")


@router.post("/import-folder")
async def import_employees_from_workspace_folder(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Import all Excel files from the workspace excel folder into the database."""
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can import employees")

    workspace_root = Path(__file__).resolve().parent.parent
    excel_dir = workspace_root / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)

    import_results = []
    errors = []
    processed_files = []

    for path in sorted(excel_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
            workbook_rows = _read_workbook_rows(workbook)
            imported_count = 0
            for sheet_name, rows in workbook_rows:
                for row_data in rows:
                    try:
                        if not row_data.get("file_code"):
                            continue
                        _upsert_employee_from_mapping(db, row_data)
                        imported_count += 1
                    except Exception as exc:  # pragma: no cover - defensive logging
                        errors.append(f"{path.name}/{sheet_name}: {exc}")
                        logger.error("Error importing row from %s sheet %s: %s", path.name, sheet_name, exc)
            db.commit()
            processed_files.append(path.name)
            import_results.append({"file": path.name, "imported": imported_count, "sheets": [name for name, _ in workbook_rows]})
        except Exception as exc:  # pragma: no cover - defensive logging
            errors.append(f"{path.name}: {exc}")
            logger.error("Error importing workbook %s: %s", path.name, exc)

    return {
        "success": True,
        "message": f"Imported data from {len(import_results)} workbook(s)",
        "files": processed_files,
        "results": import_results,
        "errors": errors,
    }


@router.get("/employee-template")
async def get_employee_import_template(current_user: models.User = Depends(get_current_user)):
    """
    Get Excel template for employee import.
    Returns a sample Excel file with required and optional columns.
    """
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can access templates")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"

    headers = [
        "file_code",
        "full_name",
        "email",
        "position",
        "contact_number",
        "project",
        "location",
        "employment_type",
        "date_of_appointment",
        "contract_start",
        "contract_end",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E5A96", end_color="1E5A96", fill_type="solid")

    sample_data = [
        "EMP001",
        "John Doe",
        "john@example.com",
        "Senior Manager",
        "+256-123-456789",
        "Project Alpha",
        "Kampala",
        "Permanent",
        "2023-01-15",
        "2023-01-15",
        "2026-01-14",
    ]

    for col_idx, value in enumerate(sample_data, 1):
        worksheet.cell(row=2, column=col_idx).value = value

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return {
        "message": "Template created. Use columns: file_code (required), full_name (required), email (required), position (required), and optional columns for other data.",
        "required_columns": ["file_code", "full_name", "email", "position"],
        "optional_columns": ["contact_number", "project", "location", "employment_type", "date_of_appointment", "contract_start", "contract_end"],
    }

