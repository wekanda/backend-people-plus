"""
Document Forms - faithful, print-ready official document templates.

Every template in this module reproduces the corresponding master document from the
`word documents` folder (TPO Uganda). They are rendered as A4 HTML so a user can fill
a form in the app, print it, and get output that looks exactly like the original file.
"""

import html as _html
import re
import io
import os
from datetime import datetime

###############################################################################
# Shared A4 print stylesheet mirrors the Word layout (justified text, 1.5 line
# height, Proxima Nova / Calibri style typeface, bold + underline headings,
# dotted signature lines and bordered tables).
###############################################################################
PAGE_CSS = """
html, body { margin: 0; padding: 0; background: #fff; }
.doc-page {
  width: 210mm;
  min-height: 297mm;
  margin: 0 auto;
  padding: 22mm 24mm;
  box-sizing: border-box;
  background: #fff;
  color: #000;
  font-family: "Proxima Nova", "Segoe UI", Calibri, Arial, sans-serif;
  font-size: 12pt;
  line-height: 1.5;
  text-align: justify;
  -webkit-print-color-adjust: exact;
  print-color-adjust: exact;
}
.doc-page .center { text-align: center; }
.doc-page .right { text-align: right; }
.doc-page .bold { font-weight: 700; }
.doc-page .ul { text-decoration: underline; }
.doc-page h1 { font-size: 16pt; text-align: center; margin: 8pt 0 14pt; font-weight: 700; }
.doc-page h2 { font-size: 13.5pt; font-weight: 700; margin: 16pt 0 8pt; }
.doc-page h3 { font-size: 12.5pt; font-weight: 700; margin: 12pt 0 6pt; }
.doc-page p { margin: 0 0 6pt; }
.doc-page ul { margin: 0 0 6pt 18pt; padding: 0; }
.doc-page li { margin: 0 0 4pt; }
.doc-page .blank { height: 1em; }
.doc-page table { border-collapse: collapse; width: 100%; margin: 8pt 0; }
.doc-page table.bordered th, .doc-page table.bordered td {
  border: 1px solid #000; padding: 5pt 7pt; vertical-align: top;
}
.doc-page table.info td { border: none; padding: 3pt 5pt; vertical-align: top; }
.doc-page table.info td.k { font-weight: 700; width: 44%; }
.doc-page .field-value {
  display: inline-block;
  min-width: 7em;
  border-bottom: 0;
  font-weight: 400;
}
.doc-page .empty-field {
  display: inline-block;
  min-width: 7em;
  border-bottom: 1px dotted #444;
}
.doc-page .sig { margin-top: 10pt; }
.doc-page .spacer-xl { height: 20mm; }
.doc-page .spacer-lg { height: 14mm; }
.doc-page .spacer-md { height: 9mm; }
.doc-page .spacer-sm { height: 4mm; }
.doc-page .dots {
  font-family: "Courier New", monospace;
  letter-spacing: 2px;
  white-space: nowrap;
}
@media print {
  body * { visibility: hidden; }
  .doc-page, .doc-page * { visibility: visible; }
  .doc-page {
    position: absolute;
    left: 0; top: 0;
    width: 210mm;
    margin: 0;
    box-shadow: none;
  }
  @page { size: A4; margin: 0; }
}
.doc-page { box-shadow: 0 2px 20px rgba(0,0,0,0.15); }
"""

DEFAULT_SIGN = "……………………………………………………………………………"

def esc(t):
    return _html.escape(str(t), quote=True)

def fmt_date(v):
    """Convert '2026-02-01' -> '1st February 2026' (matches Word masters)."""
    if v in (None, ""):
        return ""
    v = str(v).strip()
    if not v:
        return ""
    d = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            d = datetime.strptime(v, fmt)
            break
        except ValueError:
            continue
    if d is None:
        return v
    day = d.day
    suffix = "th" if (4 <= day <= 20 or 24 <= day <= 30) else {1: "st", 2: "nd", 3: "rd"}[day % 10]
    return f"{day}{suffix} {d.strftime('%B %Y')}"

def fmt_money(v):
    """Format numbers with thousands separators: 2500977 -> 2,500,977."""
    if v in (None, ""):
        return ""
    try:
        f = float(str(v).replace(",", "").replace(" ", ""))
        return f"{f:,.0f}"
    except Exception:
        return str(v)

def h(v, placeholder=""):
    """Render a value (escaped) or an empty dotted line when blank."""
    if v is None:
        v = ""
    v = str(v)
    if str(v).strip():
        return f'<span class="field-value">{esc(str(v).strip())}</span>'
    return f'<span class="empty-field">{esc(placeholder)}</span>'

def hd(v, placeholder=""):
    return h(fmt_date(v), placeholder)

def hm(v, placeholder=""):
    return h(fmt_money(v), placeholder)
# Regex for template placeholders:
#   {h(date)}  {h(name, 'placeholder')}  {hm(money)}  {hd(date)}  {first_name(name)}
_H_RE = re.compile(r"\{h\(([\w_]+)\)\}")
_HP_RE = re.compile(r"\{h\(([\w_]+),\s*'(.*?)'\)\}")
_HD_RE = re.compile(r"\{hd\(([\w_]+)\)\}")
_HM_RE = re.compile(r"\{hm\(([\w_]+)\)\}")
_FN_RE = re.compile(r"\{first_name\(([\w_]+)\)\}")


def _apply_template(tpl, ctx):
    """Substitute {h(x)} / {h(x,'ph')} / {hd(x)} / {hm(x)} / {first_name(x)}."""
    def repl_fn(m):
        name = m.group(1)
        val = ctx.get(name, "")
        fn = str(val).strip().split()[0] if str(val).strip() else "____"
        return esc(fn)

    def repl_h(m):
        return h(ctx.get(m.group(1), ""))

    def repl_hp(m):
        return h(ctx.get(m.group(1), ""), m.group(2))

    def repl_hd(m):
        return hd(ctx.get(m.group(1), ""))

    def repl_hm(m):
        return hm(ctx.get(m.group(1), ""))

    tpl = _FN_RE.sub(repl_fn, tpl)
    tpl = _HD_RE.sub(repl_hd, tpl)
    tpl = _HM_RE.sub(repl_hm, tpl)
    tpl = _HP_RE.sub(repl_hp, tpl)
    tpl = _H_RE.sub(repl_h, tpl)
    tpl = tpl.replace("{sign_dots}", DEFAULT_SIGN)
    return tpl


# Map Employee database columns to form fields (single source of truth -> no retyping).
EMPLOYEE_FIELD_MAP = {
    "full_name": ["full_name"],
    "employee_name": ["full_name"],
    "to_name": ["full_name"],
    "intern_name": ["full_name"],
    "contractor_name": ["full_name"],
    "employee_no": ["file_code"],
    "employee_id": ["file_code"],
    "position": ["position"],
    "to_title": ["position"],
    "employee_title": ["position"],
    "job_title": ["position"],
    "contractor_title": ["position"],
    "department": ["project"],
    "project": ["project"],
    "project_name": ["project"],
    "duty_station": ["location"],
    "location": ["location"],
    "supervisor": ["supervisor"],
    "report_to": ["supervisor"],
    "phone": ["contact_number"],
    "principal_contact": ["contact_number"],
    "email": ["email"],
    "principal_email": ["email"],
    "date_of_appointment": ["date_of_appointment"],
    "appointment_date": ["date_of_appointment"],
    "contract_start": ["contract_start"],
    "new_start": ["contract_start"],
    "contract_end": ["contract_end"],
    "new_end": ["contract_end"],
    "expiry_date": ["contract_end"],
    "employment_type": ["employment_type"],
    "notice_period": ["notice_period"],
    "status": ["status"],
    "grade": ["grade"],
}

def autofill_from_employee(employee):
    """Build fill-in values for the Forms Library from an employee database record."""
    source = {}
    for attr in ("full_name", "file_code", "position", "project", "location",
                 "contact_number", "date_of_appointment", "contract_start",
                 "contract_end", "employment_type", "notice_period", "status", "grade"):
        v = getattr(employee, attr, None)
        if v is None:
            continue
        source[attr] = v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else v
    values_map = {}
    for fname, attrs in EMPLOYEE_FIELD_MAP.items():
        for a in attrs:
            if source.get(a) not in (None, ""):
                values_map[fname] = source[a]
                break
    return values_map

def autofill_forms_from_employee(employee):
    """Return previews + values for all forms given an employee DB record."""
    common = autofill_from_employee(employee)
    results = {}
    for form in DOCUMENT_FORMS:
        allowed = {f["name"] for f in form.get("fields", [])}
        values = {k: v for k, v in common.items() if k in allowed}
        results[form["key"]] = {
            "name": form["name"],
            "values": values,
            "preview": render_document_html(form, values),
        }
    return results


def render_document_html(form, values=None):
    """Render a document form dict (or key) to a complete A4 HTML page."""
    if isinstance(form, str):
        form = get_form(form)
    if form is None:
        raise KeyError("Unknown form")
    ctx = dict(values or {})
    for f in form.get("fields", []):
        name = f["name"]
        raw = ctx.get(name, "")
        if raw in (None, ""):
            ctx[name] = f.get("default", "")
        else:
            ftype = f.get("type", "text")
            if ftype == "date":
                ctx[name] = fmt_date(raw)
            elif ftype == "number":
                ctx[name] = fmt_money(raw)
    # compute procurement total when not explicitly provided
    if form.get("key") == "procurement_request" and ctx.get("item_total", "") in (None, ""):
        try:
            def _num(v):
                if v in (None, ""):
                    return 0.0
                return float(str(v).replace(",", "").strip())
            tot = sum(_num(ctx.get(k)) for k in ("item1_cost", "item2_cost", "item3_cost"))
            ctx["item_total"] = tot or ""
        except Exception:
            ctx["item_total"] = ""
    # Leave tracker balances = entitled - taken (auto-computed)
    if form.get("key") == "leave_tracker":
        for base in ("annual", "sick", "maternity", "paternity"):
            en, tk = ctx.get(f"{base}_entitled", ""), ctx.get(f"{base}_taken", "")
            try:
                bal = (float(str(en).replace(",", "")) if en not in (None, "") else 0) - \
                      (float(str(tk).replace(",", "")) if tk not in (None, "") else 0)
                ctx[f"{base}_balance"] = bal if bal else ""
            except Exception:
                ctx[f"{base}_balance"] = ""
    body = _apply_template(form["template"], ctx)
    title = esc(form.get("name", "Document"))
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<style>{PAGE_CSS}</style>
</head>
<body>
<div class="doc-page">
{body}
</div>
</body>
</html>"""


def get_form(key):
    for f in DOCUMENT_FORMS:
        if f["key"] == key:
            return f
    return None


# Official legal letters/contracts may only be GENERATED by HR Admin, per the
# role-based design guide (Manager & Staff cannot generate contracts or letters).
RESTRICTED_GENERATION = {
    "offer_letter", "employment_contract", "contract_extension", "contract_renewal",
    "end_of_contract_notice", "short_term_contract", "sla", "letter_of_undertaking",
    "practicum_placement",
}
ALL_GENERATE_ROLES = ("hr_admin", "project_manager", "staff", "finance")

def generate_roles_for(form):
    return ["hr_admin"] if form["key"] in RESTRICTED_GENERATION else list(ALL_GENERATE_ROLES)

def can_generate(user_role, form):
    return user_role in generate_roles_for(form)


def get_field_map(form):
    return {f["name"]: f for f in form.get("fields", [])}


# Common column names users are likely to have in their Excel files.
FIELD_ALIASES = {
    "employee_name": ["full name", "employee name", "name of employee", "staff name", "name"],
    "to_name": ["full name", "employee name", "name", "applicant name", "staff name", "intern name", "contractor name"],
    "intern_name": ["name of intern", "intern", "student name"],
    "contractor_name": ["name of contractor", "contractor ", "vendor name", "contractor name"],
    "applicant_name": ["name of applicant", "applicant", "borrower name"],
    "to_title": ["position", "job title", "job position", "designation", "role"],
    "employee_title": ["position", "job title", "job position", "designation", "role"],
    "position": ["job title", "job position", "designation", "role"],
    "contractor_title": ["title", "role"],
    "supervisor": ["reporting to", "line manager", "manager", "supervisor name"],
    "report_to": ["reports to", "supervisor"],
    "project": ["project name", "project title", "project"],
    "project_name": ["project", "project title", "programme"],
    "date": ["letter date", "document date", "generation date", "date"],
    "contract_start": ["appointment date", "commencement date", "start date", "contract start date"],
    "contract_end": ["contract end date", "expiry date", "termination date", "end date"],
    "start_date": ["commencement date", "start date"],
    "end_date": ["completion date", "end date"],
    "salary_gross": ["gross salary", "gross pay", "monthly salary", "basic salary"],
    "gross_pay": ["gross", "gross salary", "gross pay"],
    "net_pay": ["net salary", "net pay", "net"],
    "duty_station": ["station", "duty station"],
    "location": ["location", "station"],
    "duty_station": ["station", "location"],
    "overall_rating": ["rating", "performance rating"],
    "duration": ["duration", "term"],
    "amount_figures": ["amount", "loan amount"],
    "monthly_installment": ["monthly instalment", "instalment"],
    "effective_date": ["effective date", "extension start"],
    "expiry_date": ["expiry date", "current expiry date"],
}

def _field_candidates(fdef):
    name = fdef["name"]
    cands = {normalize(name), normalize(fdef.get("label", ""))}
    for alias in FIELD_ALIASES.get(name, []):
        cands.add(normalize(alias))
    cands.discard("")
    return sorted(cands, key=len, reverse=True)

def normalize(text):
    return re.sub(r"\s+", " ", str(text).lower().replace("_", " ").replace("-", " ")).strip()


def excel_autofill(filename, file_bytes, mapping=None):
    """Parse an uploaded Excel workbook and derive fill-in values for every form.

    mapping (optional): {form_key: {field_name: column_header}}.
    When absent, columns are matched to fields by label/name similarity.
    Returns: {filename, sheet, rows, columns, forms: {key: {name, values, preview}}, unmatched_forms:[...]}
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"filename": filename, "forms": {}, "sheet": ws.title, "rows": 0}

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    records = []
    for r in rows[1:]:
        rec = {}
        for i, hd_ in enumerate(headers):
            if hd_ and i < len(r):
                rec[hd_] = r[i]
        if any(str(v).strip() for v in rec.values()):
            records.append(rec)

    norm_headers = {}
    for hd_ in headers:
        if not hd_:
            continue
        norm = hd_.lower().replace("_", " ").replace("-", " ").replace("  ", " ").strip()
        norm_headers.setdefault(norm, hd_)
        norm_headers.setdefault(hd_.lower(), hd_)

    def first_value(col):
        for rec in records:
            if rec.get(col) not in (None, ""):
                return rec[col]
        return ""

    results = {}
    unmatched = []
    for form in DOCUMENT_FORMS:
        fm = get_field_map(form)
        values_map = {}
        if mapping and form["key"] in mapping:
            for fname, col in mapping[form["key"]].items():
                if col in norm_headers.values() or col in headers:
                    values_map[fname] = first_value(col)
        # Prefer the most specific header first (longest) so a field like
        # contract_start picks "contract start date" instead of plain "date".
        ordered_hk = sorted(norm_headers.keys(), key=len, reverse=True)
        for fname, fdef in fm.items():
            if fname in values_map:
                continue
            cands = _field_candidates(fdef)
            col = None
            # 1) exact match on a candidate
            for cand in cands:
                if cand in norm_headers:
                    col = norm_headers[cand]
                    break
            # 2) containment: a header fully contains the candidate
            if col is None:
                for cand in cands:
                    for hk in ordered_hk:
                        if cand in hk:
                            col = norm_headers[hk]
                            break
                    if col:
                        break
            if col:
                values_map[fname] = first_value(col)

        if not values_map:
            unmatched.append(form["name"])
        else:
            results[form["key"]] = {
                "name": form["name"],
                "values": values_map,
                "preview": render_document_html(form, values_map),
            }
    return {
        "filename": filename,
        "sheet": ws.title,
        "rows": len(records),
        "columns": [c for c in headers if c],
        "forms": results,
        "unmatched_forms": unmatched,
    }


DOCUMENT_FORMS = []

def _load_registry():
    from . import form_templates
    for name in sorted(dir(form_templates)):
        if name.startswith("_"):
            continue
        obj = getattr(form_templates, name)
        if isinstance(obj, dict) and obj.get("key"):
            DOCUMENT_FORMS.append(obj)

_load_registry()
if not DOCUMENT_FORMS:
    raise RuntimeError("No document forms registered under backend.form_templates")