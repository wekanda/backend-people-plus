"""
Excel Import Router
Handles bulk employee imports from Excel and CSV files.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime, date
from database import get_db
from auth import get_current_user
import models
import openpyxl
import csv
from io import BytesIO, StringIO
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/excel", tags=["excel"])


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


@router.post("/import-employees")
async def import_employees_from_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Bulk import employees from Excel or CSV file.
    Expected columns:
    - file_code (required)
    - full_name (required)
    - email (required)
    - position (required)
    - contact_number
    - project
    - location
    - employment_type
    - date_of_appointment
    - contract_start
    - contract_end
    """
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can import employees")

    try:
        file_content = await file.read()
        file_name = (file.filename or "").lower()

        if file_name.endswith((".xlsx", ".xls")):
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(min_row=2, values_only=False))
            headers = {cell.value.lower().strip(): idx + 1 for idx, cell in enumerate(worksheet[1]) if cell.value}

            def get_cell(row, key):
                if key not in headers:
                    return None
                return row[headers[key] - 1].value
        elif file_name.endswith(".csv"):
            text = file_content.decode("utf-8", errors="ignore")
            csv_reader = csv.DictReader(StringIO(text))
            if not csv_reader.fieldnames:
                raise HTTPException(status_code=400, detail="CSV file has no headers")
            headers = {header.lower().strip(): header for header in csv_reader.fieldnames}
            rows = list(csv_reader)

            def get_cell(row, key):
                if key not in headers:
                    return None
                return row.get(headers[key])
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload .xlsx, .xls, or .csv")

        required_fields = ["file_code", "full_name", "email", "position"]
        for field in required_fields:
            if field not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {field}")

        imported_employees = []
        errors = []
        total_rows = 0

        for row_idx, row in enumerate(rows, start=2):
            try:
                total_rows += 1
                file_code = get_cell(row, "file_code")
                full_name = get_cell(row, "full_name")
                email = get_cell(row, "email")
                position = get_cell(row, "position")

                if not all([file_code, full_name, email, position]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue

                contact_number = get_cell(row, "contact_number") or None
                project = get_cell(row, "project") or None
                location = get_cell(row, "location") or None
                employment_type = get_cell(row, "employment_type") or "Contract"
                date_of_appointment = _parse_date(get_cell(row, "date_of_appointment") or "")
                contract_start = _parse_date(get_cell(row, "contract_start") or "")
                contract_end = _parse_date(get_cell(row, "contract_end") or "")

                if contact_number is not None:
                    contact_number = str(contact_number)

                existing = db.query(models.Employee).filter(models.Employee.file_code == str(file_code)).first()

                if existing:
                    existing.full_name = full_name
                    existing.position = position
                    existing.contact_number = contact_number
                    existing.project = project
                    existing.location = location
                    existing.employment_type = employment_type
                    if date_of_appointment:
                        existing.date_of_appointment = date_of_appointment
                    if contract_start:
                        existing.contract_start = contract_start
                    if contract_end:
                        existing.contract_end = contract_end
                    existing.status = "Active"
                    db.add(existing)
                    imported_employees.append({"file_code": file_code, "name": full_name, "status": "updated"})
                else:
                    employee = models.Employee(
                        file_code=str(file_code),
                        full_name=full_name,
                        position=position,
                        contact_number=contact_number,
                        project=project,
                        location=location,
                        employment_type=employment_type,
                        date_of_appointment=date_of_appointment,
                        contract_start=contract_start,
                        contract_end=contract_end,
                        status="Active"
                    )
                    db.add(employee)
                    imported_employees.append({"file_code": file_code, "name": full_name, "status": "created"})
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                logger.error(f"Error importing row {row_idx}: {str(e)}")
                db.rollback()
                continue

        db.commit()
        return {
            "success": True,
            "message": f"Imported {len(imported_employees)} employees",
            "imported": imported_employees,
            "errors": errors,
            "total_rows_processed": total_rows,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error importing file: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error importing file: {str(e)}")


@router.get("/employee-template")
async def get_employee_import_template(current_user: models.User = Depends(get_current_user)):
    """
    Get an Excel template for employee import.
    """
    if current_user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admins can access templates")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"

    headers = [
        "file_code",
        "full_name",
        "email",
        "position",
        "contact_number",
        "project",
        "location",
        "employment_type",
        "date_of_appointment",
        "contract_start",
        "contract_end",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="1E5A96", end_color="1E5A96", fill_type="solid")

    sample_data = [
        "EMP001",
        "John Doe",
        "john@example.com",
        "Senior Manager",
        "+256-123-456789",
        "Project Alpha",
        "Kampala",
        "Permanent",
        "2023-01-15",
        "2023-01-15",
        "2026-01-14",
    ]

    for col_idx, value in enumerate(sample_data, 1):
        worksheet.cell(row=2, column=col_idx).value = value

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return {
        "message": "Template created. Use columns: file_code (required), full_name (required), email (required), position (required), and optional columns for other data.",
        "required_columns": ["file_code", "full_name", "email", "position"],
        "optional_columns": ["contact_number", "project", "location", "employment_type", "date_of_appointment", "contract_start", "contract_end"],
    }
