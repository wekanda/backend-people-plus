"""
Document Forms Router - list, render print-ready HTML, download, save and Excel autofill.
"""
import io
import json

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
import models
from backend import form_documents

router = APIRouter(prefix="/api/form-documents", tags=["form-documents"])

ALLOWED_ROLES = ["hr_admin", "project_manager", "staff", "finance"]


def _check_role(user):
    if user.role not in ALLOWED_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return user


@router.get("")
@router.get("/")
def list_form_documents(user=Depends(get_current_user)):
    """Return metadata + typed fields for every form document."""
    _check_role(user)
    return {
        "count": len(form_documents.DOCUMENT_FORMS),
        "forms": [_public_form(f) for f in form_documents.DOCUMENT_FORMS],
    }


@router.get("/saved")
def list_saved_forms(db: Session = Depends(get_db), user=Depends(get_current_user)):
    """List the current user's saved form drafts."""
    _check_role(user)
    rows = db.query(models.SavedForm).filter(models.SavedForm.created_by == user.id).order_by(
        models.SavedForm.updated_at.desc()).all()
    return {
        "saved": [{
            "id": r.id,
            "form_key": r.form_key,
            "form_name": r.form_name,
            "employee_id": r.employee_id,
            "values": json.loads(r.values_json) if r.values_json else {},
            "preview": r.html or "",
            "saved_at": r.updated_at.isoformat() if r.updated_at else None,
        } for r in rows]
    }


@router.delete("/saved/{saved_id}")
def delete_saved_form(saved_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Delete one of the user's saved form drafts."""
    _check_role(user)
    row = db.query(models.SavedForm).filter(models.SavedForm.id == saved_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Saved form not found")
    if row.created_by != user.id and user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Cannot delete another user's saved form")
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.get("/{key}")
def get_form_document(key: str, user=Depends(get_current_user)):
    """Return metadata + fields for one form document."""
    _check_role(user)
    form = form_documents.get_form(key)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    return _public_form(form)


@router.post("/{key}/render")
def render_form_document(key: str, payload: dict, user=Depends(get_current_user)):
    """Render a filled document to print-ready A4 HTML."""
    _check_role(user)
    form = form_documents.get_form(key)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    if not form_documents.can_generate(user.role, form):
        raise HTTPException(status_code=403, detail="Only HR Admin can generate contractual documents")
    html = form_documents.render_document_html(form, payload.get("values") or {})
    return {"key": key, "name": form["name"], "html": html}


@router.post("/{key}/download")
def download_form_document(key: str, payload: dict, user=Depends(get_current_user)):
    """Return the filled document as a printable Word-compatible .doc file."""
    _check_role(user)
    form = form_documents.get_form(key)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    if not form_documents.can_generate(user.role, form):
        raise HTTPException(status_code=403, detail="Only HR Admin can download official documents")
    html = form_documents.render_document_html(form, payload.get("values") or {})
    word_html = html.replace(
        "<html",
        '<html xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:w="urn:schemas-microsoft-com:office:word" '
        'xmlns="http://www.w3.org/TR/REC-html40"')
    word_html = word_html.replace(
        "<head>",
        '<head><xml><w:WordDocument><w:View>Print</w:View>'
        '<w:Zoom>100</w:Zoom></w:WordDocument></xml>')
    filename = f"{key}.doc"
    return Response(content=word_html, media_type="application/msword",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.post("/excel-autofill")
async def excel_autofill(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Upload an Excel file; its columns are mapped onto every document form (HR Admin)."""
    _check_role(user)
    if user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admin can autofill document templates")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    try:
        result = form_documents.excel_autofill(file.filename or "upload.xlsx", data)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {exc}")
    return result


@router.get("/employee-autofill/{employee_id}")
def employee_autofill(employee_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Auto-populate all document forms from an employee record (HR Admin)."""
    _check_role(user)
    if user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admin can generate documents from employee records")
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    results = form_documents.autofill_forms_from_employee(employee)
    return {
        "employee": {
            "id": employee.id,
            "full_name": employee.full_name,
            "file_code": employee.file_code,
            "position": employee.position,
            "project": employee.project,
        },
        "forms": results,
    }


@router.post("/excel-template")
async def download_excel_template(user=Depends(get_current_user)):
    """Return a reusable Excel template whose headers match the form fields (HR Admin)."""
    _check_role(user)
    if user.role != "hr_admin":
        raise HTTPException(status_code=403, detail="Only HR Admin can download the Excel field template")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Document Fields"
    ws.append(["FORM", "FIELD NAME", "FIELD LABEL", "DATA TYPE", "EXAMPLE"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1D4ED8")
        c.alignment = Alignment(horizontal="center")
    examples = {
        "text": "e.g. Jane Atim",
        "date": "2026-01-01",
        "number": "2500977",
        "longtext": "Paragraph describing duties...",
        "select": "Option A / Option B",
    }
    for f in form_documents.DOCUMENT_FORMS:
        for field in f["fields"]:
            ws.append([f["name"], field["name"], field.get("label", ""),
                       field.get("type", "text"), examples.get(field.get("type", "text"), "")])
    for col, width in zip("ABCDE", (34, 24, 32, 12, 36)):
        ws.column_dimensions[col].width = width
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="form_documents_template.xlsx"'})


@router.post("/{key}/save")
def save_form_document(key: str, payload: dict, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Persist the filled values for the current user (draft storage)."""
    _check_role(user)
    form = form_documents.get_form(key)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    if not form_documents.can_generate(user.role, form):
        raise HTTPException(status_code=403, detail="Only HR Admin can generate contractual documents")
    values = payload.get("values") or {}
    html = form_documents.render_document_html(form, values)
    employee_id = payload.get("employee_id")
    saved = db.query(models.SavedForm).filter(
        models.SavedForm.form_key == key,
        models.SavedForm.created_by == user.id,
    ).first()
    if saved:
        saved.values_json = json.dumps(values, ensure_ascii=False)
        saved.html = html
        saved.employee_id = employee_id
        saved.form_name = form["name"]
    else:
        saved = models.SavedForm(
            form_key=key,
            form_name=form["name"],
            employee_id=employee_id,
            values_json=json.dumps(values, ensure_ascii=False),
            html=html,
            created_by=user.id,
        )
        db.add(saved)
    db.commit()
    db.refresh(saved)
    return {"ok": True, "id": saved.id, "form_key": key, "saved_at": saved.updated_at.isoformat()}


def _public_form(form):
    fields = []
    for f in form.get("fields", []):
        fields.append({
            "name": f["name"],
            "label": f.get("label", f["name"]),
            "type": f.get("type", "text"),
            "required": bool(f.get("required", False)),
            "default": f.get("default", ""),
            "placeholder": f.get("placeholder", ""),
            "options": f.get("options", []),
        })
    return {
        "key": form["key"],
        "name": form["name"],
        "category": form.get("category", ""),
        "description": form.get("description", ""),
        "fields": fields,
        "generate_roles": form_documents.generate_roles_for(form),
    }
    fields = []
    for f in form.get("fields", []):
        fields.append({
            "name": f["name"],
            "label": f.get("label", f["name"]),
            "type": f.get("type", "text"),
            "required": bool(f.get("required", False)),
            "default": f.get("default", ""),
            "placeholder": f.get("placeholder", ""),
            "options": f.get("options", []),
        })
    return {
        "key": form["key"],
        "name": form["name"],
        "category": form.get("category", ""),
        "description": form.get("description", ""),
        "fields": fields,
    }