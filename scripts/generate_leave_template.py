from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

wb = Workbook()
ws = wb.active
ws.title = 'Leave Application'

# Column widths approximate
for col, width in [('A', 18), ('B', 40), ('C', 18), ('D', 40)]:
    ws.column_dimensions[col].width = width

# Title
ws.merge_cells('A1:D1')
ws['A1'] = 'LEAVE APPLICATION FORM'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# General Information
ws['A3'] = 'NAME OF EMPLOYEE'
ws['B3'] = ''
ws['A4'] = 'EMPLOYEE ID NO'
ws['B4'] = ''
ws['A5'] = 'DESIGNATION/TITLE'
ws['B5'] = ''
ws['A6'] = 'DATE OF APPOINTMENT'
ws['B6'] = ''

# Leave Duration
ws['A8'] = 'NO OF DAYS TO BE TAKEN'
ws['B8'] = ''
ws['A9'] = 'FROM'
ws['B9'] = ''
ws['A10'] = 'TO'
ws['B10'] = ''
ws['A11'] = 'RETURN DATE'
ws['B11'] = ''

# Leave Type
ws['A13'] = 'TYPE OF LEAVE'
ws['B13'] = ''

# Computation area
ws['A15'] = 'COMPUTATION OF LEAVE'
ws['B15'] = ''

out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates', 'excel', 'Leave_Application_template.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)

print('Generated', out_path)
